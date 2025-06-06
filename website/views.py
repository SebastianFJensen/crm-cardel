from django.shortcuts import render , redirect, HttpResponse, get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.models import User
from django.contrib import messages
from .models import Record , Comment , RecordResource, Folder, File, get_file_location
from .forms import AddRecordForms , CommentForm, ImportRecordDataForm, NewFolderForm
from django.views.generic.edit import CreateView, UpdateView
from django.views.generic import CreateView
from django.db.models import Case, When, Value, F, DateField, CharField
from django.urls import reverse_lazy
from .resources import RecordResource
from django.contrib.auth.decorators import login_required
from tablib import Dataset
import pandas as pd
from rest_framework import generics
from rest_framework.parsers import MultiPartParser
from import_export.formats.base_formats import CSV , XLS , XLSX
from import_export import resources
from django.views.generic import View
from openpyxl import load_workbook
from .filters import RecordFilter
from django.utils import timezone
from datetime import datetime
import unicodedata
from azure.storage.blob import BlobServiceClient
from django.conf import settings
import os



def home(request):
    # Always fetch records (for unauthenticated users and staff)
    ordering = Case(
        When(Forfaldsdato__isnull=True, then=Value('9999-12-31')),
        default=F('Forfaldsdato'),
        output_field=DateField()
    )
    records = Record.objects.order_by(ordering)

    # Handle login for unauthenticated users
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")
            if user.is_staff:
                return render(request, 'home.html', {'records': records})
            else:
                return redirect('sommerhuse')  # Redirect non-staff after login
        else:
            messages.error(request, "Ugyldigt brugernavn eller adgangskode")

    # For unauthenticated users or staff, render home.html
    return render(request, 'home.html', {'records': records})

def search(request):
    query = request.GET.get('q', '')
    records = Record.objects.filter(Adresse__icontains=query) | Record.objects.filter(BFE_Nummer__icontains=query) | Record.objects.filter(Telefonnummer__icontains=query) | Record.objects.filter(By__icontains=query)

    if request.user.is_authenticated:
        return render(request, 'search.html', {'records': records})
    else:
        messages.success(request, "Du skal være logget ind for at se siden")
        return redirect('home')


def archived(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")

    ordering = Case(
    When(Opfølgningsdato__isnull=True, then=Value('9999-12-31')),
    default=F('Opfølgningsdato'),
    output_field=DateField()
    ).asc()

    records = Record.objects.order_by(ordering, '-created_at')
    return render(request, 'archived.html', {'records': records})

def prospects(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")

    ordering = Case(
    When(Opfølgningsdato__isnull=True, then=Value('9999-12-31')),
    default=F('Opfølgningsdato'),
    output_field=DateField()
    ).asc()

    records = Record.objects.order_by(ordering, '-created_at')
    return render(request, 'prospects.html', {'records': records})

def lead(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")

    ordering = Case(
    When(Opfølgningsdato__isnull=True, then=Value('9999-12-31')),
    default=F('Opfølgningsdato'),
    output_field=DateField()
    ).asc()

    records = Record.objects.order_by(ordering, '-created_at')
    return render(request, 'lead.html', {'records': records})

def sommerhuse(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, "Du er logget ind")

    ordering = Case(
    When(Opfølgningsdato__isnull=True, then=Value('9999-12-31')),
    default=F('Opfølgningsdato'),
    output_field=DateField()
    ).asc()

    records = Record.objects.order_by(ordering, '-created_at')
    return render(request, 'sommerhuse.html', {'records': records})


def login_user (request): 
	pass

def logout_user (request):
	logout(request)
	messages.success(request, "Du er nu logget ud")
	return redirect('home')

class CustomerRecordView(View):
    def get(self, request, pk):
        if request.user.is_authenticated:
            customer_record = Record.objects.get(id=pk)
            folders = customer_record.folders.all()
            return render(request, "Record.html", {'customer_record':customer_record, 'folders':folders})
        else:
            messages.success(request, "Du skal være logget ind for at se siden")
            return redirect('home')

    def post(self, request, pk):
        if request.user.is_authenticated:
            customer_record = Record.objects.get(id=pk)
            messages.success(request, "Sagen er blevet gemt")
            return redirect('Record')
        else:
            messages.success(request, "Du skal være logget ind for at se siden")
            return redirect('home')

def login_user(request): 
    pass

def logout_user(request):
    logout(request)
    messages.success(request, "Du er nu logget ud")
    return redirect('home')


def delete_record(request, pk):
    if not request.user.is_authenticated:
        messages.error(request, "Du skal være logget ind for at se siden")
        return redirect('home')
    if not request.user.is_staff:
        messages.error(request, "Du har ikke rettigheder til dette område")
        return redirect('home')

    record = get_object_or_404(Record, pk=pk)
    record.delete()
    messages.success(request, "Sagen er blevet slettet")
    return redirect('home')

def add_record(request):
    if not request.user.is_authenticated:
        messages.error(request, "Du skal være logget ind for at se siden")
        return redirect('home')
    if not request.user.is_staff:
        messages.error(request, "Du har ikke rettigheder til dette område")
        return redirect('home')

    form = AddRecordForms(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            bfe_nummer = form.cleaned_data['BFE_Nummer']
            if not Record.objects.filter(BFE_Nummer=bfe_nummer).exists():
                add_record = form.save()
                messages.success(request, "Registrering er tilføjet...")
                return redirect('Record', pk=add_record.pk)
            else:
                messages.error(request, "Registreringen findes allerede...")
                return render(request, 'add_record.html', {'form':form})
    return render(request, 'add_record.html', {'form':form})


def update_record(request, pk):
    if not request.user.is_authenticated:
        messages.error(request, "Du skal være logget ind for at se siden")
        return redirect('home')
    if not request.user.is_staff:
        messages.error(request, "Du har ikke rettigheder til dette område")
        return redirect('home')

    customer_record = Record.objects.get(id=pk)
    form = AddRecordForms(request.POST or None, instance=customer_record)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Sagen er blevet opdateret")
            return redirect('Record', pk=customer_record.pk)
    return render(request, 'update_record.html', {'customer_record':customer_record, 'form':form, 'pk':pk})

def record_details(request, pk):
    record = Record.objects.get(id=pk)
    return render(request, 'Record.html', {'record': record})


class AddCommentView(CreateView):
    model = Comment
    form_class = CommentForm
    template_name = 'add_comment.html'

    def form_valid(self, form):
        form.instance.post_id = self.kwargs['pk']
        form.instance.user = self.request.user
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('Record', kwargs={'pk': self.kwargs['pk']})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

class ImportRecordData(View):
    form_class = ImportRecordDataForm
    template_name = 'importrecorddata.html'

    def get(self, request):
        form = self.form_class(None)
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            # import record data logic goes here
            # ...
            return redirect('home')
        return render(request, self.template_name, {'form': form})


def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            created_at, BFE_Nummer, Adresse, Kommune, Region, Kontaktperson, Mail, Telefonnummer, m2, Kommuneplan, Lokalplan, Formaal = row
            customer_record.objects.create(created_at=created_at, BFE_Nummer=BFE_Nummer, Adresse=Adresse, Kommune=Kommune, Region=Region, Kontaktperson=Kontaktperson, Mail=Mail, Telefonnummer=Telefonnummer, m2=m2, Kommuneplan=Kommuneplan, Lokalplan=Lokalplan, Formål=Formål)

        return render(request, 'import_success.html')

    return render(request, 'import_form.html')

#def sendt_til_dla(request, pk)
	#customer_record = customer_record.objects.get(id=pk)
	#customer_record.sendt_til_dla == True if request.GET.get('Sendt_til_DLA') == 'True' else False
	#customer_record.save()

def create_new_folder(request):
    if request.method == 'POST':
        form = NewFolderForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponse('Mappe oprettet')
    else:
        createfolder = NewFolderForm
    return render(request, 'createnewfolder.html', {'createfolder':createfolder})

def open_folder(request, pk):
    folder = get_object_or_404(Folder, pk=pk)
    case_id = folder.record.id
    request.session['case_id'] = case_id
    return render(request, 'openfolder.html', {'folder':folder})

def upload_file(request):
    if not request.user.is_authenticated:
        messages.error(request, "Du skal være logget ind for at se siden")
        return redirect('home')
    if not request.user.is_staff:
        messages.error(request, "Du har ikke rettigheder til dette område")
        return redirect('home')

    if request.method == 'POST':
        folder_id = request.POST.get('fid', '')
        if not folder_id:
            messages.error(request, "Folder ID is required")
            return redirect('open_folder', pk=folder_id)

        folder = get_object_or_404(Folder, pk=folder_id)

        for uploaded_file in request.FILES.getlist('uploadfile[]'):
            if len(uploaded_file.name) > 300:
                messages.error(request, "Filnavn er for langt (maks. 300 tegn): %s" % uploaded_file.name)
                return redirect('open_folder', pk=folder_id)

            # Create a BlobServiceClient object
            blob_service_client = BlobServiceClient.from_connection_string(settings.AZURE_CONNECTION_STRING)

            # Create a blob client
            filename_base = unicodedata.normalize('NFKD', uploaded_file.name).encode('ASCII', 'ignore').decode()
            blob_client = blob_service_client.get_blob_client("cardel", f"{folder.record.id}/{folder.folder_type}/{filename_base}")

            # Upload the file
            blob_client.upload_blob(uploaded_file, overwrite=True)

            # Get the URL of the uploaded file
            file_url = blob_client.url

            # Check if a file with the same name already exists in the database
            existing_file = File.objects.filter(folder=folder, files=uploaded_file.name).first()
            if existing_file:
                # Update the existing file
                existing_file.file_url = file_url
                existing_file.uploaded_on = datetime.now()
                existing_file.save()
            else:
                # Create a new File instance
                new_file = File(folder=folder, files=uploaded_file.name, user=request.user, file_url=file_url, uploaded_on=datetime.now())
                new_file.save()

    return redirect(reverse_lazy('Record', kwargs={'pk': folder.record.pk}))
    
def delete_file(request, pk):
    if not request.user.is_authenticated:
        messages.error(request, "Du skal være logget ind for at se siden")
        return redirect('home')
    if not request.user.is_staff:
        messages.error(request, "Du har ikke rettigheder til dette område")
        return redirect('home')

    file = get_object_or_404(File, pk=pk)

    # Create a BlobServiceClient object
    blob_service_client = BlobServiceClient.from_connection_string(settings.AZURE_CONNECTION_STRING)

    # Generate the file path using the get_file_location function
    file_path = get_file_location(file, file.files.name)

    # Create a blob client
    blob_client = blob_service_client.get_blob_client("cardel", file_path)

    # Delete the blob
    blob_client.delete_blob()

    # Delete the file instance
    file.delete()

    messages.success(request, "Filen er blevet slettet")
    return redirect('open_folder', pk=file.folder.pk)

class EditCommentView(UpdateView):
    model = Comment
    form_class = CommentForm
    template_name = 'edit_comment.html'

    def get_object(self, queryset=None):
        comment = get_object_or_404(Comment, id=self.kwargs['comment_id'], user=self.request.user)
        return comment

    def form_valid(self, form):
        form.instance.modified_on = timezone.now()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('Record', kwargs={'pk': self.object.post.id})
    
